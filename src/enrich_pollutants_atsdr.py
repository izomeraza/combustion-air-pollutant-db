#!/usr/bin/env python3
"""Join CAS-deduped pollutants with the ATSDR Substance Priority List.

Input:
  outputs/candidate_combustion_pollutants_by_cas_pubchem.csv

Output:
  outputs/candidate_combustion_pollutants_by_cas_pubchem.csv

Matching is exact on normalized CASRN only.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = Path("outputs/candidate_combustion_pollutants_by_cas_pubchem.csv")
DEFAULT_ATSDR = Path("ATSDR-2025-Official-SPL.xlsx")
DEFAULT_OUTPUT = Path("outputs/candidate_combustion_pollutants_by_cas_pubchem.csv")
LEGACY_FIELDS = {
    "atsdr_profile_titles",
    "atsdr_profile_urls",
    "atsdr_pdf_urls",
    "atsdr_brief_health_hazard_summary",
    "atsdr_mrl_summary",
}


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_csv_atomic(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with tmp_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    tmp_path.replace(path)


def normalize_cas(value: str) -> str:
    return re.sub(r"[^0-9]", "", (value or "").strip())


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def append_missing(fieldnames: list[str], candidates: list[str]) -> list[str]:
    out = list(fieldnames)
    for candidate in candidates:
        if candidate not in out:
            out.append(candidate)
    return out


def format_cell(value: object) -> str:
    return normalize_text(value)


def read_atsdr_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "SPL Data" not in workbook.sheetnames:
        raise SystemExit("ATSDR workbook missing 'SPL Data' sheet")
    sheet = workbook["SPL Data"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [format_cell(cell) for cell in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        record = {header[i]: format_cell(row[i]) if i < len(row) else "" for i in range(len(header))}
        out.append(record)
    return out


def build_atsdr_index(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    by_cas: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        cas = normalize_cas(row.get("CASRN", ""))
        if cas:
            by_cas[cas].append(row)
    return by_cas


def merge_atsdr_fields(rows: list[dict[str, str]]) -> dict[str, str]:
    if not rows:
        return {
            "atsdr_spl_year": "",
            "atsdr_spl_rank": "",
            "atsdr_spl_rank_previous_year": "",
            "atsdr_spl_rank_change": "",
            "atsdr_substance_name": "",
            "atsdr_casrn": "",
            "atsdr_total_points": "",
            "atsdr_toxicity_weight": "",
            "atsdr_toxicity_source": "",
            "atsdr_toxicity_points": "",
            "atsdr_total_exposure_points": "",
            "atsdr_match_count": "0",
        }

    def unique_join(key: str) -> str:
        values: list[str] = []
        for row in rows:
            value = normalize_text(row.get(key, ""))
            if value and value not in values:
                values.append(value)
        return " | ".join(values)

    return {
        "atsdr_spl_year": unique_join("Year"),
        "atsdr_spl_rank": unique_join("Rank"),
        "atsdr_spl_rank_previous_year": unique_join("Rank from Previous Year"),
        "atsdr_spl_rank_change": unique_join("Rank Change from Previous Year"),
        "atsdr_substance_name": unique_join("Substance Name"),
        "atsdr_casrn": unique_join("CASRN"),
        "atsdr_total_points": unique_join("Total Points"),
        "atsdr_toxicity_weight": unique_join("Toxicity Weight"),
        "atsdr_toxicity_source": unique_join("Toxicity Source"),
        "atsdr_toxicity_points": unique_join("Toxicity Points"),
        "atsdr_total_exposure_points": unique_join("Total Exposure Points"),
        "atsdr_match_count": str(len(rows)),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Join combustion pollutants with ATSDR SPL data.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--atsdr", type=Path, default=DEFAULT_ATSDR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")
    if not args.atsdr.exists():
        raise SystemExit(f"ATSDR workbook not found: {args.atsdr}")

    input_rows = read_csv_dicts(args.input)
    atsdr_rows = read_atsdr_rows(args.atsdr)
    atsdr_index = build_atsdr_index(atsdr_rows)

    cleaned_input_rows: list[dict[str, str]] = []
    for row in input_rows:
        cleaned = {key: value for key, value in row.items() if key not in LEGACY_FIELDS}
        cleaned_input_rows.append(cleaned)

    fieldnames = append_missing(
        list(cleaned_input_rows[0].keys()),
        [
            "atsdr_spl_year",
            "atsdr_spl_rank",
            "atsdr_spl_rank_previous_year",
            "atsdr_spl_rank_change",
            "atsdr_substance_name",
            "atsdr_casrn",
            "atsdr_total_points",
            "atsdr_toxicity_weight",
            "atsdr_toxicity_source",
            "atsdr_toxicity_points",
            "atsdr_total_exposure_points",
            "atsdr_match_count",
        ],
    )

    joined_rows: list[dict[str, str]] = []
    matched_rows = 0
    matched_cas: set[str] = set()
    for row in cleaned_input_rows:
        cas = normalize_cas(row.get("cas", ""))
        records = atsdr_index.get(cas, [])
        enriched = dict(row)
        enriched.update(merge_atsdr_fields(records))
        joined_rows.append(enriched)
        if records:
            matched_rows += 1
            matched_cas.add(cas)

    write_csv_atomic(args.output, joined_rows, fieldnames)

    total = len(cleaned_input_rows)
    print(f"wrote {len(joined_rows)} rows to {args.output}")
    print(f"matched rows: {matched_rows}/{total} ({matched_rows / total:.1%})")
    print(f"matched unique CAS: {len(matched_cas)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
