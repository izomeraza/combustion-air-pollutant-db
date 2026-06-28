#!/usr/bin/env python3
"""Join pollutants with EPA IRIS regulatory toxicity values.

Input:
  outputs/candidate_combustion_pollutants_by_cas_pubchem_iarc.csv

Output:
  outputs/candidate_combustion_pollutants_by_cas_pubchem.csv

Matching is exact on normalized CASRN only.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = Path("outputs/candidate_combustion_pollutants_by_cas_pubchem_iarc.csv")
DEFAULT_IRIS = Path("simple_list_alpha.xlsx")
DEFAULT_OUTPUT = Path("outputs/candidate_combustion_pollutants_by_cas_pubchem.csv")


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_csv_atomic(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with tmp_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    tmp_path.replace(path)


def normalize_cas(value: str) -> str:
    return re.sub(r"[^0-9]", "", (value or "").strip())


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return normalize_text(str(value))


def read_iris_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [format_cell(cell) for cell in rows[0]]
    iris_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {header[i]: format_cell(row[i]) if i < len(row) else "" for i in range(len(header))}
        iris_rows.append(record)
    return iris_rows


def unique_join(records: list[dict[str, str]], key: str) -> str:
    values: list[str] = []
    for record in records:
        value = normalize_text(record.get(key, ""))
        if value and value not in values:
            values.append(value)
    return " | ".join(values)


def append_missing(fieldnames: list[str], candidates: list[str]) -> list[str]:
    out = list(fieldnames)
    for candidate in candidates:
        if candidate not in out:
            out.append(candidate)
    return out


def merge_iris_records(records: list[dict[str, str]]) -> dict[str, str]:
    if not records:
        return {
            "iris_chemical_name": "",
            "iris_last_significant_revision": "",
            "iris_literature_screening_review": "",
            "iris_critical_effect_systems": "",
            "iris_rfd_chronic": "",
            "iris_rfd_subchronic": "",
            "iris_rfc_chronic": "",
            "iris_rfc_subchronic": "",
            "iris_tumor_site": "",
            "iris_match_count": "0",
        }

    return {
        "iris_chemical_name": unique_join(records, "Chemical Name"),
        "iris_last_significant_revision": unique_join(records, "Last Significant Revision"),
        "iris_literature_screening_review": unique_join(records, "Literature Screening Review"),
        "iris_critical_effect_systems": unique_join(records, "Critical Effect Systems"),
        "iris_rfd_chronic": unique_join(records, "RfD (Chronic)"),
        "iris_rfd_subchronic": unique_join(records, "RfD (Subchronic)"),
        "iris_rfc_chronic": unique_join(records, "RfC (Chronic)"),
        "iris_rfc_subchronic": unique_join(records, "RfC (Subchronic)"),
        "iris_tumor_site": unique_join(records, "Tumor Site"),
        "iris_match_count": str(len(records)),
    }


def build_iris_index(iris_rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    by_cas: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in iris_rows:
        cas = normalize_cas(row.get("CASRN", ""))
        if cas:
            by_cas[cas].append(row)
    return by_cas


def main() -> int:
    parser = argparse.ArgumentParser(description="Join pollutants with EPA IRIS toxicity values.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--iris", type=Path, default=DEFAULT_IRIS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")
    if not args.iris.exists():
        raise SystemExit(f"IRIS workbook not found: {args.iris}")

    work_rows = read_csv_dicts(args.input)
    iris_rows = read_iris_rows(args.iris)
    iris_index = build_iris_index(iris_rows)

    joined_rows: list[dict[str, str]] = []
    matched_rows = 0
    matched_cas: set[str] = set()
    for row in work_rows:
        cas = normalize_cas(row.get("cas", ""))
        records = iris_index.get(cas, [])
        enriched = dict(row)
        enriched.update(merge_iris_records(records))
        joined_rows.append(enriched)
        if records:
            matched_rows += 1
            matched_cas.add(cas)

    fieldnames = append_missing(
        list(work_rows[0].keys()),
        [
            "iris_chemical_name",
            "iris_last_significant_revision",
            "iris_literature_screening_review",
            "iris_critical_effect_systems",
            "iris_rfd_chronic",
            "iris_rfd_subchronic",
            "iris_rfc_chronic",
            "iris_rfc_subchronic",
            "iris_tumor_site",
            "iris_match_count",
        ],
    )
    write_csv_atomic(args.output, joined_rows, fieldnames)

    total = len(work_rows)
    print(f"wrote {len(joined_rows)} rows to {args.output}")
    print(f"matched rows: {matched_rows}/{total} ({matched_rows / total:.1%})")
    print(f"matched unique CAS: {len(matched_cas)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
